import os
from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook

app = Flask(__name__)

EXCEL_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "data",
    "EduBridge_Database_Data.xlsx"
)

COLUMNS = [
    "Name",
    "Address",
    "PinCode",
    "Contact",
    "Website",
    "Category",
    "Amount"
]


def load_data():
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True
    )

    sheet = workbook["trust"]
    rows = []

    for values in sheet.iter_rows(min_row=2, values_only=True):

        if not any(value is not None for value in values):
            continue

        item = {}

        for index, column in enumerate(COLUMNS):
            value = values[index] if index < len(values) else None
            item[column] = "" if value is None else str(value).strip()

        rows.append(item)

    workbook.close()
    return rows


def get_categories(rows):
    categories = set()

    for row in rows:
        category = row.get("Category", "").strip()

        if not category:
            continue

        main_category = category.split(",")[0].strip()

        if main_category:
            categories.add(main_category)

    return sorted(categories, key=str.lower)


@app.route("/")
def index():

    rows = load_data()
    categories = get_categories(rows)

    return render_template(
        "index.html",
        scholarships=rows,
        categories=categories
    )


@app.route("/search")
def search():

    pincode = request.args.get("pincode", "").strip()
    category = request.args.get("category", "").strip()

    rows = load_data()
    results = []

    for row in rows:

        row_category = row.get("Category", "")
        row_pincode = row.get("PinCode", "")

        # Category filter
        if category and category != "All":

            if category.lower() not in row_category.lower():
                continue

        # PinCode filter
        if pincode:

            if not (
                len(pincode) == 6
                and pincode.isdigit()
            ):
                continue

            if pincode not in row_pincode:
                continue

        results.append(row)

    # Exact PinCode matches first
    if pincode and len(pincode) == 6 and pincode.isdigit():

        results.sort(
            key=lambda item:
            0 if item.get("PinCode") == pincode else 1
        )

    return jsonify(results)


@app.route("/api/data")
def api_data():
    return jsonify(load_data())


@app.route("/health")
def health():

    return {
        "status": "ok",
        "database": "Excel",
        "file": "EduBridge_Database_Data.xlsx"
    }


if __name__ == "__main__":
    app.run(debug=True)
